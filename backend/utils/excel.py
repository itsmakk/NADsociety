"""
NADSOC — Excel Generation Utilities
On-demand Excel export (NOT stored on server). SRS Section 2.
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# NADSOC Blue theme colors
HEADER_FILL = PatternFill(start_color='1A56DB', end_color='1A56DB', fill_type='solid')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
DATA_FONT = Font(name='Arial', size=10)
TITLE_FONT = Font(name='Arial', bold=True, size=14, color='0D2B72')
SUBTITLE_FONT = Font(name='Arial', bold=True, size=11, color='374151')
BORDER = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)


def create_excel_report(title, subtitle, columns, data, sheet_name='Report'):
    """
    Generate an Excel workbook with standard NADSOC formatting.

    Args:
        title: Report title (e.g., 'DEMAND/RECOVERY STATEMENT')
        subtitle: Subtitle (e.g., 'MARCH 2026 | INDUSTRIAL / PART-A')
        columns: List of {'header': str, 'key': str, 'width': int, 'format': str}
        data: List of dicts matching column keys
        sheet_name: Worksheet name

    Returns:
        BytesIO buffer containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # --- Title Row ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    title_cell = ws.cell(row=1, column=1, value='THE NAD EMPLOYEES CO-OP. CREDIT SOCIETY LTD, KARANJA')
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal='center')

    # --- Subtitle Row ---
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    sub_cell = ws.cell(row=2, column=1, value=f'{title} — {subtitle}')
    sub_cell.font = SUBTITLE_FONT
    sub_cell.alignment = Alignment(horizontal='center')

    # --- Empty Row ---
    start_row = 4

    # --- Header Row ---
    for col_idx, col in enumerate(columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=col['header'])
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col.get('width', 15)

    # --- Data Rows ---
    for row_idx, row_data in enumerate(data, start_row + 1):
        for col_idx, col in enumerate(columns, 1):
            value = row_data.get(col['key'], '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = BORDER

            fmt = col.get('format', 'text')
            if fmt == 'number':
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0.00'
            elif fmt == 'integer':
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0'
            elif fmt == 'date':
                cell.number_format = 'DD/MM/YYYY'
            else:
                cell.alignment = Alignment(horizontal='left')

    # --- Print Setup ---
    ws.print_title_rows = f'1:{start_row}'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # Write to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def create_member_upload_template():
    """Generate the bulk member upload Excel template (SRS 4.1.1)."""
    columns = [
        {'header': 'GEN No *', 'key': 'gen_no', 'width': 12},
        {'header': 'Name *', 'key': 'name', 'width': 25},
        {'header': 'Token No', 'key': 'token_no', 'width': 12},
        {'header': 'Designation', 'key': 'designation', 'width': 18},
        {'header': 'Employee Type *', 'key': 'employee_type', 'width': 18},
        {'header': 'Part', 'key': 'part', 'width': 12},
        {'header': 'DOJ (DD/MM/YYYY)', 'key': 'doj', 'width': 16},
        {'header': 'Status', 'key': 'status', 'width': 12},
        {'header': 'LL Bal', 'key': 'll_bal', 'width': 14, 'format': 'number'},
        {'header': 'LLP', 'key': 'llp', 'width': 14, 'format': 'number'},
        {'header': 'SL Bal', 'key': 'sl_bal', 'width': 14, 'format': 'number'},
        {'header': 'SLP', 'key': 'slp', 'width': 14, 'format': 'number'},
        {'header': 'CD Bal', 'key': 'cd_bal', 'width': 14, 'format': 'number'},
        {'header': 'Share Bal', 'key': 'share_bal', 'width': 14, 'format': 'number'},
        {'header': 'Email', 'key': 'email', 'width': 25},
        {'header': 'Mobile No', 'key': 'mobile', 'width': 14},
        {'header': 'Bank Name', 'key': 'bank_name', 'width': 20},
        {'header': 'Account No', 'key': 'account_no', 'width': 18},
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Members'

    # Instructions sheet
    ws_help = wb.create_sheet('Instructions')
    instructions = [
        'NADSOC — Bulk Member Upload Template',
        '',
        'RULES:',
        '1. Fields marked with * are mandatory',
        '2. GEN No must be unique',
        '3. Employee Type: Industrial or Non-Industrial',
        '4. Date format: DD/MM/YYYY',
        '5. LLP ≤ LL Bal (Interest = Balance - Principal)',
        '6. SLP ≤ SL Bal',
        '7. No negative values allowed',
        '8. The ENTIRE file is rejected if even one row has errors',
        '',
        'STATUS VALUES: Active, Inactive',
        '',
        'If email is not provided, system generates GENNO@nadsoc.local',
        'Default password will be DOB (DDMMYYYY)',
    ]
    for i, line in enumerate(instructions, 1):
        ws_help.cell(row=i, column=1, value=line).font = DATA_FONT

    # Headers
    for col_idx, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col['header'])
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col['width']

    # Sample row
    sample = {
        'gen_no': '1001', 'name': 'Sample Member', 'token_no': 'T001',
        'designation': 'Officer', 'employee_type': 'Industrial', 'part': 'Part-A',
        'doj': '01/04/2020', 'status': 'Active',
        'll_bal': 100000, 'llp': 80000, 'sl_bal': 20000, 'slp': 18000,
        'cd_bal': 50000, 'share_bal': 8000,
        'email': 'sample@example.com', 'mobile': '9876543210',
        'bank_name': 'SBI', 'account_no': '12345678901'
    }
    for col_idx, col in enumerate(columns, 1):
        ws.cell(row=2, column=col_idx, value=sample.get(col['key'], ''))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
