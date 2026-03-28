"""
NADSOC — Bulk Upload API Routes (SRS 4.1.1)
Excel-based member + opening balance upload.
"""

import json
from datetime import datetime
from flask import Blueprint, request, jsonify, send_file
from middleware.auth import require_auth, get_current_username
from middleware.rbac import require_permission
from utils.db import get_transaction, set_app_user
from utils.excel import create_member_upload_template
from models.schemas import sanitize_string
from services.audit_service import log_action

upload_bp = Blueprint('upload', __name__)


@upload_bp.route('/template', methods=['GET'])
@require_auth
@require_permission('members', 'bulk_upload')
def download_template():
    """GET /api/upload/template — Download bulk upload Excel template."""
    buffer = create_member_upload_template()
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='NADSOC_Member_Upload_Template.xlsx'
    )


@upload_bp.route('/members', methods=['POST'])
@require_auth
@require_permission('members', 'bulk_upload')
def upload_members():
    """
    POST /api/upload/members — Bulk upload members with opening balances.
    SRS 4.1.1: Atomic transaction — entire file rejected if any row invalid.
    """
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'Only .xlsx files are accepted'}), 400

    try:
        from openpyxl import load_workbook
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'error': f'Failed to read Excel file: {str(e)}'}), 400

    # Parse header row
    headers = [str(cell.value or '').strip().lower() for cell in ws[1]]
    required_headers = ['gen no', 'name', 'employee type']
    missing_headers = [h for h in required_headers if h not in [x.replace('*', '').strip() for x in headers]]
    if missing_headers:
        return jsonify({'error': f'Missing required columns: {", ".join(missing_headers)}'}), 400

    # Header mapping
    col_map = {}
    header_aliases = {
        'gen no': 'gen_no', 'gen no *': 'gen_no', 'name': 'name', 'name *': 'name',
        'token no': 'token_no', 'designation': 'designation',
        'employee type': 'employee_type', 'employee type *': 'employee_type',
        'part': 'part', 'doj (dd/mm/yyyy)': 'doj', 'status': 'status',
        'll bal': 'll_bal', 'llp': 'llp', 'sl bal': 'sl_bal', 'slp': 'slp',
        'cd bal': 'cd_bal', 'share bal': 'share_bal',
        'email': 'email', 'mobile no': 'mobile', 'bank name': 'bank_name',
        'account no': 'account_no'
    }
    for idx, h in enumerate(headers):
        key = header_aliases.get(h.replace('*', '').strip())
        if key:
            col_map[key] = idx

    # Parse data rows
    rows = []
    errors = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:  # skip empty rows
            continue

        row_data = {}
        for key, col_idx in col_map.items():
            val = row[col_idx] if col_idx < len(row) else None
            row_data[key] = val

        # Validate row
        row_errors = []

        gen_no = str(row_data.get('gen_no', '')).strip()
        if not gen_no:
            row_errors.append('GEN No is required')

        name = str(row_data.get('name', '')).strip()
        if not name:
            row_errors.append('Name is required')

        emp_type = str(row_data.get('employee_type', '')).strip()
        if emp_type not in ('Industrial', 'Non-Industrial'):
            row_errors.append('Employee Type must be Industrial or Non-Industrial')

        # Validate financial fields
        for field, label in [('ll_bal', 'LL Bal'), ('llp', 'LLP'), ('sl_bal', 'SL Bal'),
                              ('slp', 'SLP'), ('cd_bal', 'CD Bal'), ('share_bal', 'Share Bal')]:
            val = row_data.get(field)
            if val is not None and val != '':
                try:
                    val = float(val)
                    if val < 0:
                        row_errors.append(f'{label} cannot be negative')
                    row_data[field] = val
                except (ValueError, TypeError):
                    row_errors.append(f'{label} must be a number')
            else:
                row_data[field] = 0

        # LLP <= LL Bal, SLP <= SL Bal
        if float(row_data.get('llp', 0)) > float(row_data.get('ll_bal', 0)):
            row_errors.append('LLP cannot exceed LL Bal')
        if float(row_data.get('slp', 0)) > float(row_data.get('sl_bal', 0)):
            row_errors.append('SLP cannot exceed SL Bal')

        # Parse DOJ
        doj = row_data.get('doj')
        if doj:
            if isinstance(doj, datetime):
                row_data['doj'] = doj.date()
            elif isinstance(doj, str):
                try:
                    row_data['doj'] = datetime.strptime(doj, '%d/%m/%Y').date()
                except ValueError:
                    row_errors.append('DOJ must be DD/MM/YYYY')

        row_data['gen_no'] = gen_no
        row_data['name'] = name
        row_data['employee_type'] = emp_type

        if row_errors:
            errors.append({'row': row_idx, 'gen_no': gen_no, 'errors': row_errors})
        else:
            rows.append(row_data)

    # ATOMIC: Reject all if any errors
    if errors:
        return jsonify({
            'error': 'Validation failed. Fix all errors and re-upload.',
            'total_rows': len(rows) + len(errors),
            'error_count': len(errors),
            'errors': errors[:50]  # Limit response size
        }), 400

    if not rows:
        return jsonify({'error': 'No data rows found in file'}), 400

    # Process all rows in single transaction
    username = get_current_username()
    success_count = 0

    try:
        with get_transaction() as cur:
            set_app_user(cur, username)

            for row_data in rows:
                gen_no = row_data['gen_no']
                email = row_data.get('email') or f"{gen_no}@nadsoc.local"

                # Insert member
                cur.execute("""
                    INSERT INTO members (gen_no, token_no, name, designation, employee_type, part,
                                         doj, email_id, mobile_no, bank_name, bank_account_no,
                                         member_status)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    gen_no,
                    sanitize_string(row_data.get('token_no')),
                    sanitize_string(row_data['name']),
                    sanitize_string(row_data.get('designation')),
                    row_data['employee_type'],
                    sanitize_string(row_data.get('part')),
                    row_data.get('doj'),
                    email,
                    sanitize_string(row_data.get('mobile')),
                    sanitize_string(row_data.get('bank_name')),
                    sanitize_string(row_data.get('account_no')),
                    row_data.get('status', 'Active')
                ))

                # Insert deposit summary (CD + Share opening)
                cd_bal = float(row_data.get('cd_bal', 0))
                share_bal = float(row_data.get('share_bal', 0))

                cur.execute(
                    "INSERT INTO member_deposit_summary (gen_no, cd_balance, share_balance) VALUES (%s, %s, %s)",
                    (gen_no, cd_bal, share_bal)
                )

                # Opening deposit ledger entry
                if cd_bal > 0 or share_bal > 0:
                    from utils.db import execute_query
                    cur.execute("""
                        INSERT INTO deposit_ledger (gen_no, month_year, transaction_type,
                                                     cd_received, share_received, remark, posted_by)
                        VALUES (%s, %s, 'OPENING', %s, %s, 'Migration opening balance', %s)
                    """, (gen_no, '2026-03', cd_bal, share_bal, username))

                # Opening loan entries
                ll_bal = float(row_data.get('ll_bal', 0))
                llp = float(row_data.get('llp', 0))
                if ll_bal > 0:
                    ll_interest = ll_bal - llp
                    loan_id = f"LL-MIG-{gen_no}"
                    cur.execute("""
                        INSERT INTO loan_master (loan_id, member_gen_no, loan_type, loan_mode,
                            sanction_amount, disbursed_amount, interest_rate, fixed_principal_amount,
                            tenure_months, outstanding_principal, outstanding_interest, status, remark)
                        VALUES (%s, %s, 'LL', 'New', %s, %s, 11.25, 0, 0, %s, %s, 'Active', 'Migration entry')
                    """, (loan_id, gen_no, ll_bal, ll_bal, llp, ll_interest))

                    cur.execute("""
                        INSERT INTO loan_ledger (loan_id, member_gen_no, month_year, transaction_type,
                            outstanding_principal, outstanding_interest, remark, posted_by)
                        VALUES (%s, %s, '2026-03', 'OPENING', %s, %s, 'Migration opening', %s)
                    """, (loan_id, gen_no, llp, ll_interest, username))

                sl_bal = float(row_data.get('sl_bal', 0))
                slp = float(row_data.get('slp', 0))
                if sl_bal > 0:
                    sl_interest = sl_bal - slp
                    loan_id = f"SL-MIG-{gen_no}"
                    cur.execute("""
                        INSERT INTO loan_master (loan_id, member_gen_no, loan_type, loan_mode,
                            sanction_amount, disbursed_amount, interest_rate, fixed_principal_amount,
                            tenure_months, outstanding_principal, outstanding_interest, status, remark)
                        VALUES (%s, %s, 'SL', 'New', %s, %s, 11.25, 0, 0, %s, %s, 'Active', 'Migration entry')
                    """, (loan_id, gen_no, sl_bal, sl_bal, slp, sl_interest))

                    cur.execute("""
                        INSERT INTO loan_ledger (loan_id, member_gen_no, month_year, transaction_type,
                            outstanding_principal, outstanding_interest, remark, posted_by)
                        VALUES (%s, %s, '2026-03', 'OPENING', %s, %s, 'Migration opening', %s)
                    """, (loan_id, gen_no, slp, sl_interest, username))

                success_count += 1

            # Audit log for entire upload
            log_action('members', 'BULK_UPLOAD', None, {
                'total_rows': success_count,
                'uploaded_by': username
            }, cursor=cur)

    except Exception as e:
        return jsonify({'error': f'Upload failed: {str(e)}'}), 500

    return jsonify({
        'message': 'Bulk upload completed successfully',
        'success_count': success_count,
        'total_rows': success_count
    }), 201
